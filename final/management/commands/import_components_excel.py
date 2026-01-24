from django.core.management.base import BaseCommand
from ...models import Component, \
    ComponentCategory, StatusChoices  # ye dot lagake directory refernce karna mazedaar hai
# from datetime import datetime
from openpyxl import load_workbook


class Command(BaseCommand):
    help = 'Import components from Excel (trusted input)'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str)

    def handle(self, *args, **kwargs):
        excel_file_path = kwargs['excel_file']

        wb = load_workbook(excel_file_path)
        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))

            category = row_data['category']
            name = row_data['name']
            quantity = int(row_data['quantity'])

            # Get the category object
            category_obj = ComponentCategory.objects.get(comp_cate_category_name=category)


            # Create or update based on unique fields (category + name)
            working_status = StatusChoices.objects.get(status_ch_status_label="Working")
            Component.objects.update_or_create(
                comp_category=category_obj,
                comp_name=name,
                defaults={
                    'comp_quantity_available': quantity,
                    'comp_status': working_status
                }
            )
        self.stdout.write(self.style.SUCCESS("Excel import completed"))
